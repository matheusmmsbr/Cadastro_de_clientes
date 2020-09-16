from openpyxl import *
arquivo_excel = load_workbook(filename = "Registros.xlsx")
planilha1 = arquivo_excel.active

planilha1.title = "Usuarios"
planilha1.cell(row = 1, column =1, value = "Usuarios")

print("Por favor escolha uma das opções")
case = 0
x= 1
while case != 9:
    case = int(input("1 - inserir um novo registro \n2 - checar registro\n9 - fechar\nNumero: "))

    if case == 1:
        usuario = input("Insira o registro: \n")
        x = int(2)
        
        while planilha1.cell(row = int(x), column = 1).value != None:
            x = x + 1
        planilha1.cell(row = x, column= 1).value = usuario
        arquivo_excel.save("Registros.xlsx")
    elif case == 2:
        user = ""
        procura = input("Digite o usuario procurado: \n")
        while planilha1.cell(row =x, column = 1).value != "":
            if planilha1.cell(row = x, column = 1).value == procura: 
                user = planilha1.cell(row = x, column =1).value
                break
            x += 1
        if user != "":
            print("Usuario: " + str(user) + " - linha: " + str(x))
        else:
            print("Não existente")
